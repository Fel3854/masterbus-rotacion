"""Generación del archivo Santander 'PAGOS MASIVOS' para Adelantos.

Rellena el template oficial del banco (preserva logo, banner y validaciones) con
los adelantos del período. Solo se usan las columnas A–I de la pestaña 'Pagos'
(INFORMACIÓN DEL PAGO); las columnas J–N (liquidación/retención) quedan vacías.
"""

import os
from copy import copy
from datetime import date
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

TEMPLATE_PATH = os.path.join(os.path.dirname(__file__), "templates", "santander_pagos.xlsx")
SHEET = "Pagos"
START_ROW = 8           # primera fila de datos (encabezados en la fila 7)
MAX_RAZON_SOCIAL = 30   # límite del template (textLength <= 30)

# Columna -> índice (1-based). Solo se completan las de INFORMACIÓN DEL PAGO.
COL_FORMA_PAGO = 1   # A  -> "T" (Transferencia)
COL_RAZON      = 3   # C  -> apenom (truncado a 30)
COL_TIPO_DOC   = 4   # D  -> "CUIL"
COL_CUIL       = 5   # E  -> cuil (11 dígitos, texto)
COL_FECHA      = 6   # F  -> fecha de pago DD/MM/AAAA
COL_IMPORTE    = 7   # G  -> monto (numérico)
COL_CBU        = 8   # H  -> cbu (22 dígitos, texto)


def template_disponible() -> bool:
    return os.path.exists(TEMPLATE_PATH)


def generar_excel_santander(df: pd.DataFrame, fecha_pago: date) -> bytes:
    """Devuelve los bytes del .xlsx de Santander con una fila por adelanto.

    `df` debe tener las columnas: apenom, monto, cuil, cbu.
    `fecha_pago` se aplica igual a todas las filas (fecha de pago del lote).
    """
    if not template_disponible():
        raise FileNotFoundError(
            "Falta el template del banco en templates/santander_pagos.xlsx"
        )

    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[SHEET]

    # Estilo base de la primera fila de datos, para conservar bordes/fuente al
    # escribir filas nuevas más allá del rango pre-formateado del template.
    base_styles = {
        c: copy(ws.cell(row=START_ROW, column=c)._style)
        for c in (COL_FORMA_PAGO, COL_RAZON, COL_TIPO_DOC, COL_CUIL,
                  COL_FECHA, COL_IMPORTE, COL_CBU)
    }

    fecha_str = fecha_pago.strftime("%d/%m/%Y")

    for i, (_, r) in enumerate(df.iterrows()):
        fila = START_ROW + i

        def _set(col, value, num_fmt=None):
            cell = ws.cell(row=fila, column=col)
            cell._style = copy(base_styles[col])
            cell.value = value
            if num_fmt:
                cell.number_format = num_fmt
            return cell

        apenom = str(r.get("apenom", "") or "").strip()[:MAX_RAZON_SOCIAL]
        cuil = str(r.get("cuil", "") or "")
        cbu = str(r.get("cbu", "") or "")
        monto = float(r.get("monto", 0) or 0)

        _set(COL_FORMA_PAGO, "T")
        _set(COL_RAZON, apenom)
        _set(COL_TIPO_DOC, "CUIL")
        _set(COL_CUIL, cuil, num_fmt="@")
        _set(COL_FECHA, fecha_str, num_fmt="@")
        _set(COL_IMPORTE, monto, num_fmt="#,##0.00")
        _set(COL_CBU, cbu, num_fmt="@")

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
